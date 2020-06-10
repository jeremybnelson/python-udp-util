#!/usr/bin/env python
# -*- coding: utf-8 -*-

# External libs
import openpyxl

# default libs
import pathlib
import os
import codecs
from codecs import open
from pathlib import Path

# local libs
from config import ConfigSectionKey
from database import Database
from database import MSSQL
from common import expand
from common import just_file_stem


class DatabaseExcel(Database):
    """
    This extends the database.py class for the Excel Upload application
    This should be loaded with platform_audit.cfg files
    """

    # noinspection PyUnusedLocal
    def use_database_sql(self, database_name):
        command_name = 'use_database'
        sql_template = self.sql('use_database')
        sql_command = expand(sql_template)
        sql_command += '\n\n'
        self.log(command_name, sql_command)
        return sql_command

    # noinspection PyUnusedLocal
    def drop_table_sql(self, schema_name, table_name):
        command_name = 'drop_table_if_exists'
        sql_template = self.sql('drop_table_if_exists')
        sql_command = expand(sql_template)
        sql_command = f'{sql_command}\n\n'
        self.log(command_name, sql_command)
        return sql_command

    # noinspection PyUnusedLocal
    def create_table_sql(self, schema_name, table_name, workbook):
        command_name = 'create_ref_table'
        sql_template = self.sql('create_ref_table')

        # noinspection PyUnusedLocal
        column_definitions = self._column_definition_sql(table_name, workbook)

        sql_command = expand(sql_template)
        sql_command = f'{sql_command}\n\n'
        self.log(command_name, sql_command)
        return sql_command

    # noinspection PyUnusedLocal
    def create_table_sql_v2(self, schema_name, table_name, worksheet):
        command_name = 'create_ref_table'
        sql_template = self.sql('create_ref_table')

        # noinspection PyUnusedLocal
        column_definitions = self._column_definition_sql_v2(table_name, worksheet)

        sql_command = expand(sql_template)
        sql_command = f'{sql_command}\n\n'
        self.log(command_name, sql_command)
        return sql_command

    # noinspection PyUnusedLocal
    def insert_into_table_sql(self, schema_name, table_name, worksheet):
        command_name = 'insert_into_table'
        sql_template = self.sql('insert_into_table')

        # worksheet = workbook.active
        # worksheet = workbook['Data']

        column_values = self._column_values_sql(schema_name, table_name, worksheet)

        sql_command = expand(sql_template)
        sql_command = f'{sql_command}\n\n'
        sql_command.replace('values,', 'values')
        # sql_command.replace(',\n insert', '\n insert')
        return sql_command

    # noinspection PyUnusedLocal
    def create_schema_sql(self, schema_name):
        command_name = 'create_schema'
        sql_template = self.sql(command_name)
        sql_command = expand(f'{sql_template}\n\n')
        return sql_command

    def direct_execute(self, sql_command):
        try:
            cursor = self.cursor.execute(sql_command)
        except Exception as e:
            print(e)
        return cursor


    @staticmethod
    def _column_values_sql(schema_name, table_name, worksheet):
        column_values = list()

        # Todo: Hard code to first worksheet
        # worksheet = workbook.active
        # worksheet = workbook['Data']
        max_col = worksheet.max_column
        max_row = worksheet.max_row

        # append all excel defined columns to column_definitions list
        for row in worksheet.iter_rows(min_row=3, max_col=max_col, max_row=max_row):
            row_list = list()
            for cell in row:
                # if cell.value is None then empty string instead
                if cell.value is None:
                    cell.value = ''

                cell_value = str(cell.value).replace("'", "''")
                row_list.append(cell_value)

            row_joined = "('" + "', '".join(row_list) + "')"

            # Fix NULL SQL Syntax
            row_joined.replace("'NULL'", "NULL")

            # column_values.append(row_joined)
            # Append a go; statement after each 1000 lines of column values
            if len(column_values) % 1000 == 0 and len(column_values) != 0:
                # column_values.append("go")
                current_idx = len(column_values) - 1
                last_row = column_values[current_idx]
                column_values.append(f'insert into {schema_name}.{table_name} values')
                column_values.append(row_joined)
            else:
                column_values.append(row_joined)
        insert_statement = ',\n  '.join(column_values).replace('values,', 'values')

        if len(column_values) > 1000:
            # debug_list = list(range(1, len(column_values) % 1000))
            length = len(column_values)
            num = len(column_values) // 1000
            # print(debug_list)
            counter = 0
            # oh god what have i created :O
            for row_index in list(range(1, len(column_values) // 1000 + 1)):
                # row_to_be_fixed = column_values[999]

                row_index_to_be_fixed = row_index * 999 + counter
                row_to_be_fixed_dynamic = column_values[row_index_to_be_fixed]
                cleaned_insert_statement = insert_statement.replace(
                    f'{row_to_be_fixed_dynamic},', f'{row_to_be_fixed_dynamic};\n\n')
                insert_statement = cleaned_insert_statement

                counter += 1
                # cleaned_insert_statement = insert_statement.replace(f'{row_to_be_fixed},', row_to_be_fixed)

        return insert_statement

    @staticmethod
    def _bulk_column_values_sql(workbook):
        column_values = list()

        # Todo: Hard code to first worksheet
        # worksheet = workbook.active
        worksheet = workbook['Data']
        max_col = worksheet.max_column
        max_row = worksheet.max_row

        # append all excel defined columns to column_definitions list
        for row in worksheet.iter_rows(min_row=3, max_col=max_col, max_row=max_row):
            row_list = list()
            for cell in row:
                # Logic to handle escaping single quotes in cell.value
                if str(cell.value).find("'") == -1:
                    row_list.append(str(cell.value))
                else:
                    index = str(cell.value).find("'")
                    cell_value = str(cell.value)
                    new_cell_value = cell_value[:index] + "'" + cell_value[index:]
                    row_list.append(new_cell_value)

                # if cell.value is None then empty string instead
                if cell.value is None:
                    cell.value = ''

            row_joined = "union all select '" + "', '".join(row_list) + "'"

            # Fix NULL SQL Syntax
            row_joined.replace("'NULL'", "NULL")

            # Append a go; statement after each 2000 lines of column values
            if len(column_values) % 1000 == 0 and len(column_values) != 0:
                column_values.append("go")
                column_values.append("")
                column_values.append(row_joined)
            else:
                column_values.append(row_joined)

        return ',\n  '.join(column_values)

    @staticmethod
    def _column_definition_sql(table_name, workbook):
        # create a list of column specific definitions
        column_definitions = list()

        # add identity column dynamically and add to column_definitions list
        identity_column = table_name.replace('ref', '  "')
        identity_column += 'Key" bigint identity(1,1) not null primary key'
        column_definitions.append(identity_column)

        # Todo: Hard code to first worksheet
        worksheet = workbook['Data']
        # row = worksheet.iter_rows()
        max_col = worksheet.max_column

        # Append all excel defined columns to column_definitions list
        for i in range(1, max_col+1):
            header_cell = worksheet.cell(row=1, column=i)
            datatype_cell = worksheet.cell(row=2, column=i)
            column_definitions.append(f'  "{header_cell.value}" {datatype_cell.value}')

        return ',\n'.join(column_definitions)

    @staticmethod
    def _column_definition_sql_v2(table_name, worksheet):
        # create a list of column specific definitions
        column_definitions = list()

        # add identity column dynamically and add to column_definitions list
        identity_column = table_name.replace('mdm', '').replace('lkp', '').replace('ref', '')
        # identity_column = table_name.replace('lkp', '  "')
        if 'lkp' in table_name or 'Geo' in table_name:
            identity_column = f'"{identity_column}Key" int identity(1,1) not null primary key'
        else:
            identity_column = f'"{identity_column}ID" int identity(1,1) not null primary key'
        column_definitions.append(identity_column)

        # worksheet = workbook.active
        # worksheet = workbook['Data']
        # row = worksheet.iter_rows()
        max_col = worksheet.max_column


        # append all excel defined columns to column_definitions list
        for i in range(1, max_col + 1):
            header_cell = worksheet.cell(row=1, column=i)
            datatype_cell = worksheet.cell(row=2, column=i)
            column_definitions.append(f'  "{header_cell.value}" {datatype_cell.value}')

        return ',\n'.join(column_definitions)

# Everything starts here
def main():
    sdlc = 'dev'

    # ref_file_list will hold all paths to the reference definition excel files
    ref_wb_dict = {}

    #
    masterdata_directory_list = ['access', 'common', 'reference']

    # Detect and load audit reference definition files into a list
    # ToDo: make glob case insensitive
    for masterdata_directory in masterdata_directory_list:
        for ref_file in sorted(pathlib.Path(f'../ref_docs/{masterdata_directory}/').glob('*.xlsx')):
            ref_wb = openpyxl.load_workbook(ref_file, data_only=True)
            ref_table_name = just_file_stem(str(ref_file))

            # Add file name (Table Name) as key. Add workbook object as value.
            ref_wb_dict.update([(f'{masterdata_directory}.{ref_table_name}', ref_wb)])

    # Delete all output files so new ones can be generated
    for output_file in sorted(pathlib.Path('../ref_docs/ddl_output/').glob('*.sql')):
        os.remove(output_file)

    config = ConfigSectionKey('../conf', '../local')
    config.load('connect.ini')

    # ToDo: Fill in uat and prod connection names when added to connect.ini
    if sdlc == 'dev':
        connection_name = 'database:amc_dsg_udp_01_stage_dev'
    elif sdlc == 'uat':
        connection_name = 'unknown connection point'
    elif sdlc == 'prod':
        connection_name = 'database:amc_dsg_udp_01_stage_prod'
    else:
        connection_name = 'unknown connection point'

    udp_conn_config = config(connection_name)
    udp_conn_config = MSSQL(udp_conn_config)
    udp_db = DatabaseExcel('mssql_excel_upload', udp_conn_config.conn)

    for key, value in ref_wb_dict.items():
        # Instantiate DatabaseExcel object using mssql_excel_upload.cfg as platform and udp_conn_config

        sql_file = open(f"../ref_docs/ddl_output/{key}.sql", "x", encoding='utf8')

        sql_use_statement = udp_db.use_database_sql(f'udp_masterdata_{sdlc}')
        sql_drop_table = udp_db.drop_table_sql(key.split('.')[0], key.split('.')[1])
        sql_create_schema = udp_db.create_schema_sql(key.split('.')[0])

        # sql_create_table = udp_db.create_table_sql(schema_name='udp_ref', table_name=key, workbook=value)
        sql_create_table = udp_db.create_table_sql_v2(schema_name=key.split('.')[0], table_name=key.split('.')[1],
                                                      worksheet=value.worksheets[0])

        sql_file.write(sql_use_statement)
        # sql_file.write('\n begin transaction \n')
        sql_file.write(sql_create_schema)
        sql_file.write(sql_drop_table)
        sql_file.write(sql_create_table)


        # print(sql_use_statement)
        # udp_db.direct_execute(sql_use_statement)
        # print(sql_create_schema)
        # udp_db.direct_execute(sql_create_schema)
        # print(sql_drop_table)
        # udp_db.direct_execute(sql_drop_table)
        # print(sql_create_table)
        # udp_db.direct_execute(sql_create_table)
        # udp_db.direct_execute('commit')

        for sheet in [x for x in value.worksheets if x.title.lower() not in
                                                     ('documentation', 'change log', 'changelog')]:
            sql_insert_values = udp_db.insert_into_table_sql(schema_name=key.split('.')[0], table_name=key.split('.')[1]
                                                             , worksheet=sheet)
            sql_file.write(sql_insert_values)
            # print(sql_insert_values)
            # udp_db.direct_execute(sql_insert_values)

        # sql_file.write('\n end transaction \n')
        # sql_file.write('\n commit \n')

    # Clear all err_files
    for err_file in sorted(pathlib.Path(f'../ref_docs/log/').glob('*_error*')):
        os.remove(err_file)

    # Publish directly to udp_reference_<SLDC>
    for ddl_file in sorted(pathlib.Path(f'../ref_docs/ddl_output/').glob('*.sql')):
        print(f'executing {ddl_file}')
        ddl_sql = open(ddl_file, mode='r', encoding='utf8').read()
        try:
            # print(f'SQL Code: \n {ddl_sql}')
            udp_db.direct_execute(ddl_sql)
            udp_db.direct_execute('\n commit')
            print('execution successful!')
        except Exception as e:
            err_sql_file = open(f'../ref_docs/log/{ddl_file.stem}_error.sql', 'x', encoding='utf8')
            err_log_file = open(f'../ref_docs/log/{ddl_file.stem}_error.log', 'x', encoding='utf8')
            err_sql_file.write(ddl_sql)
            err_log_file.write(str(e))
            err_sql_file.close()
            err_log_file.close()


if __name__ == '__main__':
    main()
